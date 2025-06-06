import openpyxl
from openpyxl import Workbook

class ExcelFileHandler:
    @staticmethod
    def load_from_excel(file_path, class_type=None):
        """Đọc dữ liệu từ file Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            data = []
            
            # Đọc header
            headers = [cell.value for cell in sheet[1]]
            
            # Đọc dữ liệu
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):  # Bỏ qua hàng trống
                    continue
                    
                if class_type:
                    # Tạo dictionary từ dữ liệu
                    row_dict = dict(zip(headers, row))
                    
                    # Tạo đối tượng từ class_type
                    if class_type.__name__ == 'Student':
                        obj = class_type(
                            student_id=row_dict['student_id'],
                            name=row_dict['name'],
                            birth_date=row_dict['birth_date'],
                            major=row_dict['major'],
                            gender=row_dict['gender'],
                            course=row_dict['course'],
                            faculty=row_dict['faculty'],
                            class_name=row_dict['class_name']
                        )
                    elif class_type.__name__ == 'Subject':
                        obj = class_type(
                            subject_code=row_dict['subject_code'],
                            subject_name=row_dict['subject_name'],
                            credits=row_dict['credits']
                        )
                    else:
                        obj = class_type()
                        for header, value in zip(headers, row):
                            setattr(obj, header, value)
                    data.append(obj)
                else:
                    # Tạo dictionary
                    row_dict = dict(zip(headers, row))
                    data.append(row_dict)
                    
            return data
        except FileNotFoundError:
            print(f"Không tìm thấy file {file_path}")
            return []
        except Exception as e:
            print(f"Lỗi khi đọc file: {str(e)}")
            return []

    @staticmethod
    def save_to_excel(file_path, data, headers):
        """Lưu dữ liệu vào file Excel"""
        try:
            workbook = Workbook()
            sheet = workbook.active
            
            # Ghi header
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            # Ghi dữ liệu
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = getattr(item, header) if hasattr(item, header) else item.get(header)
                    sheet.cell(row=row, column=col, value=value)
            
            workbook.save(file_path)
            print(f"✓ Lưu dữ liệu thành công vào file {file_path}")
            return True
        except Exception as e:
            print(f"Lỗi khi lưu file: {str(e)}")
            return False 