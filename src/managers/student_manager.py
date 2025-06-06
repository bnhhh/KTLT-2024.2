from src.core_types.student import Student
from datetime import datetime
import openpyxl

def validate_date(date_str):
    """Validate ngày sinh"""
    if not date_str or not date_str.strip():
        return False, "Ngày sinh không được để trống"
    try:
        datetime.strptime(date_str, "%d/%m/%Y")
        return True, ""
    except ValueError:
        return False, "Ngày sinh phải có định dạng dd/mm/yyyy"

def is_valid_student_id(student_id):
    if not student_id:  # Kiểm tra nếu student_id là None hoặc rỗng
        return False
    return student_id.isdigit() and len(student_id) == 8

def is_valid_name(name):
    for char in name:
        if not ('a' <= char <= 'z' or 'A' <= char <= 'Z' or ' ' == char or 'À' <= char <= 'ỹ' or 'á' <= char <= 'ý' or 'Á' <= char <= 'Ý' or char in "ăâđêôơưĂÂĐÊÔƠƯ"):
            return False
    return True

class StudentManager:
    def __init__(self, students_data=None, students_file="students.xlsx"):
        self.students = students_data if students_data is not None else []
        self.students_file = students_file

    def validate_student_id(self, student_id, exclude_id=None):
        """Kiểm tra tính hợp lệ của mã số sinh viên"""
        if not student_id:
            return False, "MSSV không được để trống"
        if not is_valid_student_id(student_id):
            return False, "MSSV phải là 8 chữ số"
        for student in self.students:
            if str(student.student_id) == student_id and student_id != exclude_id:
                return False, "MSSV đã tồn tại"
        return True, ""

    def validate_name(self, name):
        """Kiểm tra tính hợp lệ của họ tên"""
        if not name or not name.strip():
            return False, "Họ tên không được để trống"
        if not is_valid_name(name):
            return False, "Họ tên không hợp lệ"
        return True, ""

    def validate_gender(self, gender):
        """Kiểm tra tính hợp lệ của giới tính"""
        if not gender or not gender.strip():
            return False, "Giới tính không được để trống"
        if gender.lower() not in ['nam', 'nữ']:
            return False, "Giới tính phải là 'Nam' hoặc 'Nữ'"
        return True, ""

    def validate_course(self, course):
        """Kiểm tra tính hợp lệ của khóa học"""
        if not course or not course.strip():
            return False, "Khóa học không được để trống"
        return True, ""

    def validate_faculty(self, faculty):
        """Kiểm tra tính hợp lệ của khoa viện"""
        if not faculty or not faculty.strip():
            return False, "Khoa viện không được để trống"
        return True, ""

    def validate_class_name(self, class_name):
        """Kiểm tra tính hợp lệ của lớp"""
        if not class_name or not class_name.strip():
            return False, "Lớp không được để trống"
        return True, ""

    def validate_major(self, major):
        """Kiểm tra tính hợp lệ của ngành học"""
        if not major or not major.strip():
            return False, "Ngành học không được để trống"
        return True, ""

    def find_student_by_id(self, student_id):
        """Tìm kiếm sinh viên theo mã số sinh viên"""
        for student in self.students:
            if str(student.student_id) == student_id:
                return student
        return None

    def add_student(self, student):
        """Thêm sinh viên mới vào hệ thống"""
        # Kiểm tra tính hợp lệ của tất cả các trường
        validations = [
            self.validate_student_id(student.student_id),
            self.validate_name(student.name),
            self.validate_gender(student.gender),
            (validate_date(student.birth_date)),
            self.validate_course(student.course),
            self.validate_faculty(student.faculty),
            self.validate_class_name(student.class_name),
            self.validate_major(student.major)
        ]
        
        # Kiểm tra nếu có bất kỳ trường nào không hợp lệ
        for is_valid, message in validations:
            if not is_valid:
                print(f"Lỗi: {message}")
                return False
            
        self.students.append(student)
        print("✓ Thêm sinh viên thành công!")
        return True

    def edit_student(self, student_id, new_info):
        """Cập nhật thông tin sinh viên"""
        if not student_id:
            print("MSSV không được để trống!")
            return False

        student = self.find_student_by_id(student_id)
        if not student:
            print("Không tìm thấy sinh viên!")
            return False

        # Kiểm tra mã số sinh viên mới nếu được cung cấp
        new_student_id = new_info.get('student_id')
        if new_student_id:
            is_valid_id, message = self.validate_student_id(new_student_id, exclude_id=student_id)
            if not is_valid_id:
                print(message)
                return False

        # Kiểm tra tất cả các trường khác
        for field, value in new_info.items():
            if value is not None:  # Chỉ kiểm tra các trường đang được cập nhật
                if field == 'name':
                    is_valid, message = self.validate_name(value)
                elif field == 'gender':
                    is_valid, message = self.validate_gender(value)
                elif field == 'birth_date':
                    is_valid, message = validate_date(value)
                elif field == 'course':
                    is_valid, message = self.validate_course(value)
                elif field == 'faculty':
                    is_valid, message = self.validate_faculty(value)
                elif field == 'class_name':
                    is_valid, message = self.validate_class_name(value)
                elif field == 'major':
                    is_valid, message = self.validate_major(value)
                else:
                    continue

                if not is_valid:
                    print(f"Lỗi: {message}")
                    return False

        # Cập nhật thông tin sinh viên
        for key, value in new_info.items():
            if value is not None:
                setattr(student, key, value)
        print("✓ Cập nhật thông tin thành công!")
        return True

    def delete_student(self, student_id):
        """Xóa sinh viên khỏi hệ thống"""
        initial_len = len(self.students)
        self.students = [s for s in self.students if s.student_id != student_id]
        if len(self.students) < initial_len:
            print("✓ Xóa sinh viên thành công!")
            return True
        else:
            print("Không tìm thấy sinh viên!")
            return False

    def search_students(self, search_type, keyword):
        """Tìm kiếm sinh viên theo các tiêu chí khác nhau"""
        if not keyword:
            return []
            
        keyword = keyword.strip().lower()
        results = []
        
        try:
            if search_type == '1':  # Tìm theo MSSV
                results = [s for s in self.students if keyword in str(s.student_id).lower()]
            elif search_type == '2':  # Tìm theo tên
                results = [s for s in self.students if keyword in s.name.lower()]
            elif search_type == '3':  # Tìm theo lớp
                results = [s for s in self.students if keyword in s.class_name.lower()]
                results.sort(key=lambda x: x.name.lower())
            elif search_type == '4':  # Tìm theo khoa
                results = [s for s in self.students if keyword in s.faculty.lower()]
                results.sort(key=lambda x: x.name.lower())
            else:
                print("Loại tìm kiếm không hợp lệ!")
                return []
                
            if not results:
                print("Không tìm thấy kết quả phù hợp!")
            return results
            
        except Exception as e:
            print(f"Lỗi khi tìm kiếm: {str(e)}")
            return []

    def display_student_info(self, student):
        """Hiển thị thông tin chi tiết của một sinh viên"""
        print(f"MSSV: {student.student_id}")
        print(f"Họ tên: {student.name}")
        print(f"Ngày sinh: {student.birth_date}")
        print(f"Giới tính: {student.gender}")
        print(f"Khóa học: {student.course}")
        print(f"Khoa viện: {student.faculty}")
        print(f"Lớp: {student.class_name}")
        print(f"Ngành học: {student.major}")

    def display_all_students(self, show_gpa=False):
        """Hiển thị danh sách tất cả sinh viên"""
        if not self.students:
            print("Không có sinh viên nào trong hệ thống")
            return
        print(f"\n{'='*140}")
        print(f"{'DANH SÁCH SINH VIÊN':^140}")
        print(f"{'='*140}")
        header = "{:<10} {:<25} {:<12} {:<20} {:<8} {:<15} {:<20} {:<10}".format('Mã SV', 'Họ tên', 'Ngày sinh', 'Ngành', 'Giới tính', 'Khóa học', 'Khoa viện', 'Lớp')
        if show_gpa:
            header += " {:<10} {:<5}".format('GPA', 'Loại')
        print(header)
        print(f"{'-'*140}")
        for student in self.students:
            row = "{:<10} {:<25} {:<12} {:<20} {:<8} {:<15} {:<20} {:<10}".format(student.student_id, student.name, student.birth_date, student.major, student.gender, student.course, student.faculty, student.class_name)
            if show_gpa:
                row += " {:<10.2f} {:<5}".format(student.gpa, student.grade)
            print(row)

    def load_students_from_excel(self):
        """Đọc dữ liệu sinh viên từ file Excel"""
        students = []
        try:
            workbook = openpyxl.load_workbook(self.students_file)
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(row):  # Kiểm tra nếu hàng không chứa giá trị None
                    students.append(Student.from_dict(dict(zip(header, row))))
            print("✓ Tải dữ liệu sinh viên thành công!")
        except FileNotFoundError:
            print(f"Không tìm thấy file sinh viên XLSX: {self.students_file}")
        except Exception as e:
            print(f"Lỗi khi đọc file sinh viên XLSX: {e}")
        self.students = students

    def save_students_to_excel(self):
        """Lưu dữ liệu sinh viên vào file Excel"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            header = ['student_id', 'name', 'birth_date', 'gender', 'course', 'faculty', 'class_name', 'major']
            sheet.append(header)
            for student in self.students:
                sheet.append([student.student_id, student.name, student.birth_date, student.gender, student.course, student.faculty, student.class_name, student.major])
            workbook.save(self.students_file)
            print("✓ Lưu dữ liệu sinh viên thành công!")
            return True
        except Exception as e:
            print(f"Lỗi khi lưu file sinh viên XLSX: {e}")
            return False

    def sort_by_name(self):
        """Sắp xếp danh sách sinh viên theo tên"""
        self.students.sort(key=lambda student: student.name)

    def sort_by_gpa(self):
        """Sắp xếp danh sách sinh viên theo điểm GPA"""
        self.students.sort(key=lambda student: student.gpa, reverse=True)

    def sort_by_id(self):
        """Sắp xếp danh sách sinh viên theo mã số sinh viên"""
        self.students.sort(key=lambda student: student.student_id)