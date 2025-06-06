from src.core_types.subject import Subject
import openpyxl

class SubjectManager:
    def __init__(self, subjects_data=None, subjects_file="subjects.xlsx"):
        self.subjects = subjects_data if subjects_data is not None else []
        self.subjects_file = subjects_file

    def validate_subject_code(self, subject_code):
        """Kiểm tra tính hợp lệ của mã môn học"""
        if not subject_code or not subject_code.strip():
            return False, "Mã môn học không được để trống"
        if self.find_subject_by_code(subject_code):
            return False, "Mã môn học đã tồn tại"
        return True, ""

    def validate_subject_name(self, subject_name):
        """Kiểm tra tính hợp lệ của tên môn học"""
        if not subject_name or not subject_name.strip():
            return False, "Tên môn học không được để trống"
        return True, ""

    def validate_credits(self, credits_str):
        """Kiểm tra tính hợp lệ của số tín chỉ"""
        if not credits_str or not credits_str.strip():
            return False, "Số tín chỉ không được để trống"
        try:
            credits = int(credits_str)
            if credits <= 0:
                return False, "Số tín chỉ phải là số nguyên dương"
            return True, ""
        except ValueError:
            return False, "Số tín chỉ phải là số nguyên"

    def find_subject_by_code(self, subject_code):
        """Tìm kiếm môn học theo mã môn học"""
        for subject in self.subjects:
            if subject.subject_code == subject_code:
                return subject
        return None

    def add_subject(self, subject):
        """Thêm môn học mới vào hệ thống"""
        try:
            # Kiểm tra tính hợp lệ của tất cả các trường
            validations = [
                self.validate_subject_code(subject.subject_code),
                self.validate_subject_name(subject.subject_name),
                self.validate_credits(str(subject.credits))
            ]
            
            # Kiểm tra nếu có bất kỳ trường nào không hợp lệ
            for is_valid, message in validations:
                if not is_valid:
                    print(f"Lỗi: {message}")
                    return False
                    
            self.subjects.append(subject)
            print("✓ Thêm môn học thành công!")
            return True
        except ValueError as e:
            print(f"Lỗi: {str(e)}")
            return False

    def edit_subject(self, subject_code, new_info):
        """Cập nhật thông tin môn học"""
        subject = self.find_subject_by_code(subject_code)
        if not subject:
            print("Không tìm thấy môn học!")
            return False

        try:
            # Kiểm tra các trường mới nếu được cung cấp
            if 'subject_name' in new_info:
                is_valid, message = self.validate_subject_name(new_info['subject_name'])
                if not is_valid:
                    print(f"Lỗi: {message}")
                    return False
                subject.subject_name = new_info['subject_name'].strip()

            if 'credits' in new_info:
                is_valid, message = self.validate_credits(str(new_info['credits']))
                if not is_valid:
                    print(f"Lỗi: {message}")
                    return False
                subject.credits = int(new_info['credits'])

            print("✓ Cập nhật thông tin môn học thành công!")
            return True
        except ValueError as e:
            print(f"Lỗi: {str(e)}")
            return False

    def delete_subject(self, subject_code):
        """Xóa môn học khỏi hệ thống"""
        initial_len = len(self.subjects)
        self.subjects = [s for s in self.subjects if s.subject_code != subject_code]
        if len(self.subjects) < initial_len:
            print("✓ Xóa môn học thành công!")
            return True
        else:
            print("Không tìm thấy môn học!")
            return False

    def display_all_subjects(self):
        """Hiển thị danh sách tất cả môn học"""
        if not self.subjects:
            print("Không có môn học nào trong hệ thống.")
            return
        print("\n" + "="*50)
        print("{:<15} {:<30} {:<10}".format("Mã môn học", "Tên môn học", "Tín chỉ"))
        print("="*50)
        for subject in self.subjects:
            print("{:<15} {:<30} {:<10}".format(subject.subject_code, subject.subject_name, subject.credits))
        print("="*50)

    def load_subjects_from_excel(self):
        """Đọc dữ liệu môn học từ file Excel"""
        subjects = []
        try:
            workbook = openpyxl.load_workbook(self.subjects_file)
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(row):
                    try:
                        subject_data = dict(zip(header, row))
                        subjects.append(Subject.from_dict(subject_data))
                    except ValueError as e:
                        print(f"Lỗi dữ liệu môn học: {str(e)}")
                        continue
            print("✓ Tải dữ liệu môn học thành công!")
        except FileNotFoundError:
            print(f"Không tìm thấy file môn học: {self.subjects_file}")
        except Exception as e:
            print(f"Lỗi khi đọc file môn học: {e}")
        self.subjects = subjects

    def save_subjects_to_excel(self):
        """Lưu dữ liệu môn học vào file Excel"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            header = ['subject_code', 'subject_name', 'credits']
            sheet.append(header)
            for subject in self.subjects:
                sheet.append([subject.subject_code, subject.subject_name, subject.credits])
            workbook.save(self.subjects_file)
            print("✓ Lưu dữ liệu môn học thành công!")
            return True
        except Exception as e:
            print(f"Lỗi khi lưu file môn học: {e}")
            return False